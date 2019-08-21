# ########################################################## #
#                                                            #
# Name: KEV:Constant Evaluator                               #
# Author: GGamov                                             #
# Date: 2019                                                 #
#                                                            #
# ########################################################## #

# import libraries -------------------------------------------

import pandas as pd
from openpyxl import load_workbook
import re
import io
import os

# basic input ------------------------------------------------
 
def eq_scripts_load(_sep, _subdir, _file):
    
    # if specific file selected it should be XLSX one
    if _file != "":
        
        if _subdir != '':
            _subdir = '/' + _subdir
        _subdir = '../../input' + _subdir + '/'

        _file = _subdir + _file
        
        # open excel file
        with open(_file, "rb") as f:
            inmemory_file = io.BytesIO(f.read())
        wb = load_workbook(inmemory_file, read_only = True)
        
        # read data
        r = re.compile(r'^(input\_)*stoich(iometric)*\_coefficients*$')
        st_coeff_data = pd.read_excel(_file, sheet_name = list(filter(r.search, wb.sheetnames))[0])
        
        r = re.compile(r'^(input\_)*concentrations*$')
        con_data = pd.read_excel(_file, sheet_name = list(filter(r.search, wb.sheetnames))[0], header = 1)
        
        r = re.compile(r'^(input\_)*concentrations*$')
        type_con = pd.read_excel(_file, sheet_name = list(filter(r.search, wb.sheetnames))[0]
                                 , header = None, nrows = 1).iloc[0,:]
        
        r = re.compile(r'^(input\_)*k\_constants\_log10$')
        lg_k_data = pd.read_excel(_file, sheet_name = list(filter(r.search, wb.sheetnames))[0])
        
        r = re.compile(r'^(particle|component)_names*$')
        component_name_for_yields = pd.read_excel(_file, sheet_name = list(filter(r.search, wb.sheetnames))[0]
                                                  , header = None).iat[0, 0]
        
    # use a bunch of plain text files instead
    else:
          
        if _subdir != '':
            _subdir = '/' + _subdir
        _subdir = '../../input' + _subdir + '/'

        file_names = list(os.listdir(path = _subdir))

        r = re.compile(r'^(input\_)*stoich(iometric)*\_coefficients*')
        file = list(filter(r.search, file_names))[0]
        file = _subdir + str(file)
        st_coeff_data = pd.read_csv(file, sep = _sep)
        
        r = re.compile(r'^(input\_)*k\_constants\_log10')
        file = list(filter(r.search, file_names))[0]
        file = _subdir + str(file)
        lg_k_data = pd.read_csv(file, sep = _sep, decimal = ",")
        
        r = re.compile(r'^(input\_)*concentrations*')
        file = list(filter(r.search, file_names))[0]
        file = _subdir + str(file)
        con_data = pd.read_csv(file, sep = _sep, decimal = ",", header = 1)
        
        type_con = pd.read_csv(file, sep = _sep, header = None, nrows = 1).iloc[0,:]
        
        r = re.compile(r'^(particle|component)_names*')
        file = list(filter(r.search, file_names))[0]
        file = _subdir + str(file)
        component_name_for_yields = pd.read_csv(file, header = None).iat[0, 0]

    return st_coeff_data, lg_k_data, con_data, type_con, component_name_for_yields